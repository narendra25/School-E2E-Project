from openpyxl import Workbook
from SchoolApp.pdf import render_to_pdf
from django.http import JsonResponse
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from SchoolApp.models import Student, StudentFamily, Contacts
from django.http import HttpResponse
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from xhtml2pdf import pisa
import datetime
from django.http import HttpResponse
from SchoolApp.resources import Studentresource
from tablib import Dataset
import os
import mimetypes

##Students Import Functionality
def Student_upload(request):
    if request.method == 'POST':
        person_resource = Studentresource()
        dataset = Dataset()
        new_persons = request.FILES['myfile']
        if not new_persons.name.endswith('xlsx'):
            messages.info(request, 'Wrong Format')
            return render(request, 'StudentsUpload.html')

        imported_data = dataset.load(new_persons.read(), format='xlsx')
        print(imported_data)
        for data in imported_data:
            value = Student(
                data[0],
                data[1],
                data[2],
                data[3],
                data[4],
                data[5]
            )
            value.save()
    return render(request, 'StudentsUpload.html')


#Example Download file 
def Student_Download_File(requst):
    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
# Add data to the worksheet
    ws['A1'] = 'id'
    ws['B1'] = 'rollnumber'
    ws['C1'] = 'name'
    ws['D1'] = 'email'
    ws['E1'] = 'age'
    ws['F1'] = 'gender'
# Save the workbook to a response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Student.xlsx"'
    wb.save(response)
    return response


def Student_Export_Data_To_Excel(request):
    objs = Student.objects.all()
    data = []

    for obj in objs:
        data.append({
            "id": obj.id,
            "rollnumber ": obj.rollnumber,
            "name": obj.name,
            "email": obj.email,
            "age": obj.age,
            "gender": obj.gender
        })

    # return JsonResponse({'status':200})
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="StudentData.xlsx"'
    # wb.save(response)
    pd.DataFrame(data).to_excel(response)
    return response


def StudentFamily_Export_Data_To_Excel(request):
    objs = StudentFamily.objects.all()
    data = []

    for obj in objs:
        data.append({
            "name": obj.name,
            "father_name": obj.father_name,
            "mother_name": obj.mother_name,
            "emergency": obj.emergency,
            "address": obj.address
        })

    response = HttpResponse(content_type='application/ms-csv')
    response['Content-Disposition'] ='attachment; filename="StudentFamilyData.xlsx"'
    pd.DataFrame(data).to_excel(response)
    #return JsonResponse({'status': 200})
    return response
